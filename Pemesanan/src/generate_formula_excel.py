"""
Generate Excel simulasi dengan formula yang bisa diklik per sel.
Contoh: Baby Pakcoy (SKU 805155) dan Paprika Merah (SKU 806749) di Store 1001.

Kolom INPUT (nilai): Date, Forecast_SKU, Actual_SKU, SS, Lead_Time, Review_Period,
                     Shelf_Life, Z_Scor, Stdev, Avg_Price, SL, Opening_Aging
Kolom FORMULA (klik = lihat rumus): Opening_Stock, Order, Deliver, Closing_Stock,
                                     Waste, Lost_Sale, Waste_Rp, LostSale_Rp,
                                     Sales_Rp, Waste_Pct, LostSale_Pct
"""
import pandas as pd
import numpy as np
import os
import math
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter

# ── Konfigurasi ────────────────────────────────────────────────────────────
FILE_CSV   = r"d:\program\tesis\Data-Tesis\Data-Tesis\Pemesanan\output\Solusi_A_PerStore_Filter.csv"
DIR_OUTPUT = r"d:\program\tesis\Data-Tesis\Data-Tesis\Pemesanan\output"
FILE_OUT   = os.path.join(DIR_OUTPUT, "Simulasi_Formula_Visible.xlsx")

TARGET_ITEMS = {
    805155 : "BABY PAKCOY 250G",
    806749 : "PAPRIKA MERAH",
}
TARGET_STORE = 'T-01'

# ── Warna ─────────────────────────────────────────────────────────────────
CLR_HEADER_INPUT   = "1F4E79"   # biru tua  → kolom input (nilai)
CLR_HEADER_FORMULA = "375623"   # hijau tua → kolom formula
CLR_INPUT_ROW      = "DDEEFF"   # biru muda
CLR_FORMULA_ROW    = "E2EFDA"   # hijau muda
CLR_HEADER_INFO    = "7B2C2C"   # merah tua → kolom info SKU
CLR_INFO_ROW       = "FCE4D6"   # oranye muda

# ── Helper border ─────────────────────────────────────────────────────────
thin = Side(style="thin", color="AAAAAA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr(ws, row, col, val, color_hex, font_color="FFFFFF", bold=True):
    c = ws.cell(row=row, column=col, value=val)
    c.fill = PatternFill("solid", fgColor=color_hex)
    c.font = Font(bold=bold, color=font_color, size=9)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    c.border = BORDER
    return c

def val(ws, row, col, v, fmt=None, bg=None):
    c = ws.cell(row=row, column=col, value=v)
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font = Font(size=9)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

def fml(ws, row, col, formula, fmt=None, bg=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font = Font(size=9, color="1A5276")   # biru gelap = formula
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

# ── Kolom layout ──────────────────────────────────────────────────────────
# Info (A-F): SKU, SKU DESC, Dept Name, Class, Store, Date, Day
# Input (G-S): Forecast, Actual, Avg_Price, SS, Z_Scor, Stdev,
#              Lead_Time, Review_Period, Shelf_Life, SL,
#              Opening_Aging, Note_Order, Note_Waste
# Formula (T-AD): Opening_Stock, Deliver, Order, Closing_Stock, Lost_Sale,
#                 Waste, Sales_Rp, Waste_Rp, LostSale_Rp, Waste_Pct, LostSale_Pct

INFO_COLS = [
    ("SKU",          "A"),
    ("SKU DESC",     "B"),
    ("Dept Name",    "C"),
    ("Class",        "D"),
    ("Store",        "E"),
    ("Date",         "F"),
    ("Day",          "G"),
]
INPUT_COLS = [
    ("Forecast",       "H"),
    ("Actual",         "I"),
    ("Avg Price",      "J"),
    ("SS",             "K"),
    ("Z Score",        "L"),
    ("Stdev",          "M"),
    ("Lead Time",      "N"),
    ("Review Period",  "O"),
    ("Shelf Life",     "P"),
    ("SL",             "Q"),
    ("Opening Aging",  "R"),
    ("Note Order",     "S"),
    ("Note Waste",     "T"),
    ("Minor",          "AF"),
]
FORMULA_COLS = [
    ("Opening Stock",   "U"),
    ("Deliver",         "V"),
    ("Order",           "W"),
    ("Closing Stock",   "X"),
    ("Lost Sale",       "Y"),
    ("Waste",           "Z"),
    ("Sales Rp",        "AA"),
    ("Waste Rp",        "AB"),
    ("LostSale Rp",     "AC"),
    ("Waste Pct (%)",   "AD"),
    ("LostSale Pct (%)", "AE"),
]

# kolom index integer (1-based)
def ci(letter):
    """Konversi huruf kolom Excel ke angka 1-based."""
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch.upper()) - ord('A') + 1)
    return result

# ── Baca data ─────────────────────────────────────────────────────────────
print("Membaca CSV...")
df = pd.read_csv(FILE_CSV)
df['Date'] = pd.to_datetime(df['Date'])

wb = Workbook()
wb.remove(wb.active)   # hapus sheet default

for sku_code, sku_name in TARGET_ITEMS.items():
    df_sku = df[(df['SKU'] == sku_code) &
                (df['Store'] == TARGET_STORE)].sort_values('Date').reset_index(drop=True)

    if df_sku.empty:
        # coba SKU integer vs string
        df_sku = df[(df['SKU'].astype(str) == str(sku_code)) &
                    (df['Store'].astype(str) == str(TARGET_STORE))].sort_values('Date').reset_index(drop=True)

    if df_sku.empty:
        print(f"  Data tidak ditemukan: SKU {sku_code} Store {TARGET_STORE}")
        continue

    print(f"  SKU {sku_code} ({sku_name}): {len(df_sku)} hari")

    sheet_name = f"{sku_name[:20].replace(' ','_')}_S{TARGET_STORE}"
    ws = wb.create_sheet(title=sheet_name)

    # ── Judul sheet ──────────────────────────────────────────────────────
    ws.merge_cells("A1:AF1")
    title_cell = ws["A1"]
    title_cell.value = (f"SIMULASI PEMESANAN INVENTORI - {sku_name.upper()} "
                        f"(SKU {sku_code}) | STORE {TARGET_STORE}")
    title_cell.font  = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor="1A3657")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # ── Legend baris 2 ───────────────────────────────────────────────────
    ws.merge_cells("A2:G2")
    ws["A2"].value = "Kolom INFO (abu)"
    ws["A2"].fill  = PatternFill("solid", fgColor=CLR_INFO_ROW)
    ws["A2"].font  = Font(bold=True, size=8)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("H2:T2")
    ws["H2"].value = "Kolom INPUT - nilai dari data/simulasi (biru muda)"
    ws["H2"].fill  = PatternFill("solid", fgColor=CLR_INPUT_ROW)
    ws["H2"].font  = Font(bold=True, size=8)
    ws["H2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("U2:AF2")
    ws["U2"].value = "Kolom FORMULA - klik sel untuk melihat rumus Excel (hijau muda)"
    ws["U2"].fill  = PatternFill("solid", fgColor=CLR_FORMULA_ROW)
    ws["U2"].font  = Font(bold=True, size=8)
    ws["U2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── Header baris 3 ───────────────────────────────────────────────────
    ws.row_dimensions[3].height = 30
    for label, col_letter in INFO_COLS:
        hdr(ws, 3, ci(col_letter), label, "5D6D7E", bold=True)
    for label, col_letter in INPUT_COLS:
        hdr(ws, 3, ci(col_letter), label, CLR_HEADER_INPUT)
    for label, col_letter in FORMULA_COLS:
        hdr(ws, 3, ci(col_letter), label, CLR_HEADER_FORMULA)

    # ── Isi data baris 4+ ────────────────────────────────────────────────
    for i, row_data in df_sku.iterrows():
        r = i + 4   # baris Excel (1-index, row 1=judul, 2=legend, 3=header)
        R = str(r)

        # ─ INFO columns ─
        val(ws, r, ci("A"), row_data['SKU'],        bg=CLR_INFO_ROW)
        val(ws, r, ci("B"), row_data['SKU DESC'],   bg=CLR_INFO_ROW)
        val(ws, r, ci("C"), row_data['Dept Name'],  bg=CLR_INFO_ROW)
        val(ws, r, ci("D"), row_data['Class'],      bg=CLR_INFO_ROW)
        val(ws, r, ci("E"), row_data['Store'],      bg=CLR_INFO_ROW)
        date_cell = ws.cell(row=r, column=ci("F"),
                            value=row_data['Date'].to_pydatetime())
        date_cell.number_format = "DD/MM/YYYY"
        date_cell.fill    = PatternFill("solid", fgColor=CLR_INFO_ROW)
        date_cell.border  = BORDER
        date_cell.font    = Font(size=9)
        date_cell.alignment = Alignment(horizontal="center", vertical="center")
        val(ws, r, ci("G"), row_data['Day'],        bg=CLR_INFO_ROW)

        # ─ INPUT columns ─
        val(ws, r, ci("H"), int(row_data['Forecast_SKU']),   bg=CLR_INPUT_ROW)
        val(ws, r, ci("I"), int(row_data['Actual_SKU']),     bg=CLR_INPUT_ROW)
        val(ws, r, ci("J"), round(row_data['Avg_Price'], 2), bg=CLR_INPUT_ROW,
            fmt='#,##0.00')
        val(ws, r, ci("K"), int(row_data['SS']),             bg=CLR_INPUT_ROW)
        val(ws, r, ci("L"), row_data['Z_Scor'],              bg=CLR_INPUT_ROW)
        val(ws, r, ci("M"), row_data['Stdev'],               bg=CLR_INPUT_ROW)
        val(ws, r, ci("N"), int(row_data['Lead_Time']),      bg=CLR_INPUT_ROW)
        val(ws, r, ci("O"), int(row_data['Review_Period']),  bg=CLR_INPUT_ROW)
        val(ws, r, ci("P"), int(row_data['Shelf_Life']),     bg=CLR_INPUT_ROW)
        val(ws, r, ci("Q"), row_data['SL'],                  bg=CLR_INPUT_ROW)
        # R: Opening Aging
        if r == 4:
            val(ws, r, ci("R"), round(float(row_data['Opening_Aging']), 1), bg=CLR_INPUT_ROW)
        else:
            fml(ws, r, ci("R"), f"=MAX(X{r-1}, 0)", bg=CLR_FORMULA_ROW)

        val(ws, r, ci("S"), row_data['Note_Order'],          bg=CLR_INPUT_ROW)
        val(ws, r, ci("T"), row_data['Note_Waste'],          bg=CLR_INPUT_ROW)
        val(ws, r, ci("AF"), int(row_data['Minor_MOQ']),     bg=CLR_INPUT_ROW)

        # ─ FORMULA columns ─
        # U: Opening_Stock = Opening_Aging + Deliver
        fml(ws, r, ci("U"),
            f"=R{R}+V{R}",
            bg=CLR_FORMULA_ROW)

        # V: Deliver (formula Excel murni)
        lead_time = int(row_data['Lead_Time'])
        if r < 4 + lead_time:
            fml(ws, r, ci("V"), f"=INT(H{R}*Q{R})", bg=CLR_FORMULA_ROW)
        else:
            fml(ws, r, ci("V"), f"=INT(W{r - lead_time}*Q{R})", bg=CLR_FORMULA_ROW)

        # W: Order (formula Excel murni)
        lead_time = int(row_data['Lead_Time'])
        review_period = int(row_data['Review_Period'])
        r_start = r + lead_time
        r_end = r + lead_time + review_period - 1
        closing_stock_cell = f"X{r + lead_time - 1}"
        
        if review_period == 1:
            forecast_sum_formula = f"H{r_start}"
        else:
            forecast_sum_formula = f"SUM(H{r_start}:H{r_end})"
        fml(ws, r, ci("W"), f'=IF(S{R}="Tidak",0,CEILING(MAX(0,K{R}+{forecast_sum_formula}-{closing_stock_cell}),AF{R}))', bg=CLR_FORMULA_ROW)

        # X: Closing_Stock = Opening_Stock - Actual (formula Excel murni)
        fml(ws, r, ci("X"),
            f"=MAX(U{R}-I{R}, 0)",
            bg=CLR_FORMULA_ROW)

        # Y: Lost_Sale = MAX(0, Actual - Opening_Stock)
        fml(ws, r, ci("Y"),
            f"=MAX(I{R}-U{R}, 0)",
            bg=CLR_FORMULA_ROW)

        # Z: Waste = IF(Note_Waste="Ya", ROUND(Closing_Stock*0.3, 1), 0)
        fml(ws, r, ci("Z"),
            f'=IF(T{R}="Ya", ROUND(X{R}*0.3, 1), 0)',
            bg=CLR_FORMULA_ROW)

        # AA: Sales_Rp = Actual * Avg_Price
        fml(ws, r, ci("AA"),
            f"=ROUND(I{R}*J{R}, 0)",
            bg=CLR_FORMULA_ROW, fmt='#,##0')

        # AB: Waste_Rp = Waste * Avg_Price
        fml(ws, r, ci("AB"),
            f"=ROUND(Z{R}*J{R}, 0)",
            bg=CLR_FORMULA_ROW, fmt='#,##0')

        # AC: LostSale_Rp = Lost_Sale * Avg_Price
        fml(ws, r, ci("AC"),
            f"=ROUND(Y{R}*J{R}, 0)",
            bg=CLR_FORMULA_ROW, fmt='#,##0')

        # AD: Waste_Pct = Waste_Rp / Sales_Rp * 100
        fml(ws, r, ci("AD"),
            f'=IF(AA{R}>0, ROUND(AB{R}/AA{R}*100, 2), 0)',
            bg=CLR_FORMULA_ROW, fmt='0.00"%"')

        # AE: LostSale_Pct = LostSale_Rp / Sales_Rp * 100
        fml(ws, r, ci("AE"),
            f'=IF(AA{R}>0, ROUND(AC{R}/AA{R}*100, 2), 0)',
            bg=CLR_FORMULA_ROW, fmt='0.00"%"')

    # ── Tambah baris legenda formula di bawah ────────────────────────────
    last_r = len(df_sku) + 4
    ws.merge_cells(f"A{last_r}:AF{last_r}")
    leg = ws[f"A{last_r}"]
    leg.value = (
        "KETERANGAN RUMUS: "
        "Opening_Stock = Opening_Aging + Deliver  |  "
        "Closing_Stock = MAX(Opening_Stock - Actual, 0)  |  "
        "Lost_Sale = MAX(Actual - Opening_Stock, 0)  |  "
        "Waste = IF(Note_Waste=Ya, ROUND(Closing_Stock×0.3,1), 0)  |  "
        "Waste_Pct = Waste_Rp / Sales_Rp × 100  |  "
        "LostSale_Pct = LostSale_Rp / Sales_Rp × 100"
    )
    leg.font  = Font(bold=True, size=8, color="FFFFFF")
    leg.fill  = PatternFill("solid", fgColor="1A3657")
    leg.alignment = Alignment(horizontal="left", vertical="center",
                              wrap_text=True)
    ws.row_dimensions[last_r].height = 30

    # ── Set lebar kolom ──────────────────────────────────────────────────
    col_widths = {
        "A": 9, "B": 24, "C": 14, "D": 6, "E": 7, "F": 11, "G": 5,
        "H": 9, "I": 8, "J": 11, "K": 5, "L": 7, "M": 7,
        "N": 8, "O": 9, "P": 8, "Q": 6, "R": 11, "S": 9, "T": 9,
        "U": 11, "V": 9, "W": 9, "X": 11, "Y": 9, "Z": 7,
        "AA": 13, "AB": 12, "AC": 13, "AD": 11, "AE": 13, "AF": 8,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header rows
    ws.freeze_panes = "A4"

    print(f"  Sheet '{sheet_name}' dibuat ({len(df_sku)} baris)")

wb.save(FILE_OUT)
print(f"\nFile disimpan: {FILE_OUT}")
